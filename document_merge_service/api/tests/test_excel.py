import io

import openpyxl
import pytest
from rest_framework import exceptions

from ..data import django_file
from ..engines import XlsxTemplateEngine

_structure = {
    "key0": "xdata0",
    "key1": {
        "subkey1": "xdata1",
    },
    "key2": [
        "mixed",
        "list",
        {"subkey2": "xdata2"},
    ],
}

_expect_name = ["test", "another"]


def test_render():
    tmpl = django_file("xlsx-structure.xlsx")
    engine = XlsxTemplateEngine(tmpl)
    buf = io.BytesIO()
    engine.merge(_structure, buf)
    buf.seek(0)
    doc = openpyxl.load_workbook(buf)
    for i, ws in enumerate(doc.worksheets):
        assert ws.title == _expect_name[i]
        assert ws["A1"].value == "xdata0"
        assert ws["A2"].value == "xdata1"
        assert ws["A3"].value == _expect_name[i]
        assert ws["A5"].value == "Item: mixed"
        assert ws["A6"].value == "Item: list"
        assert ws["A7"].value == "Subitem: xdata2"


@pytest.mark.parametrize(
    "available, expect_success",
    [
        ([], True),  # disabled check by empty list
        (["key0"], False),
        (["key1.subkey1"], False),
        (["key0", "key1"], False),
        (["key0", "key1.subkey1", "key2[].subkey2"], True),  # full set of vars
    ],
)
def test_validate_template(available, expect_success):
    tmpl = django_file("xlsx-structure.xlsx")
    engine = XlsxTemplateEngine(tmpl)
    if expect_success:
        engine.validate(available)
    else:
        with pytest.raises(exceptions.ValidationError):
            engine.validate(available)


def test_syntax_error():
    tmpl = django_file("xlsx-syntax.xlsx")
    engine = XlsxTemplateEngine(tmpl)
    with pytest.raises(exceptions.ValidationError):
        engine.validate(sample_data=_structure)


def test_valid_error():
    tmpl = django_file("xlsx-not-valid.xlsx")
    engine = XlsxTemplateEngine(tmpl)
    with pytest.raises(exceptions.ParseError):
        engine.validate(sample_data=_structure)
