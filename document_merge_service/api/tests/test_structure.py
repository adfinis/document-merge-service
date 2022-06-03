import io

import openpyxl

from ..data import django_file
from ..engines import XlsxTemplateEngine

_structure = {
    "key0": "xdata0",
    "key1": {
        "subkey1": "xdata1",
    },
    "key2": [
        "mixed",
        "list",
        {"subkey2": "xdata2"},
    ],
}


def test_structure():
    tmpl = django_file("xlsx-structure.xlsx")
    engine = XlsxTemplateEngine(tmpl)
    buf = io.BytesIO()
    engine.merge(_structure, buf)
    buf.seek(0)
    doc = openpyxl.load_workbook(buf)
    for ws in doc.worksheets:
        assert ws["A1"].value == "xdata0"
        assert ws["A2"].value == "xdata1"
        assert ws["A5"].value == "Item: mixed"
        assert ws["A6"].value == "Item: list"
        assert ws["A7"].value == "Subitem: xdata2"
